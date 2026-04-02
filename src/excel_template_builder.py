"""
Excel Template Builder for CR2A.

Manages the per-bid-folder workbook: copies the master Template.xlsx,
adds Contract Analysis / Specs sheets, and populates data as analyses run.

Author: CR2A Development Team
Date: 2026-03-12
"""

import copy
import logging
import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Style constants — match the CC Bid Checklist in Template.xlsx
# ---------------------------------------------------------------------------
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

_TITLE_FONT = Font(name="Times New Roman", size=12, bold=True)
_HEADER_FONT = Font(name="Times New Roman", size=12, bold=True)
_LABEL_FONT = Font(name="Times New Roman", size=12, bold=True)
_DATA_FONT = Font(name="Times New Roman", size=12, bold=False)

_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TOP_WRAP = Alignment(vertical="top", wrap_text=True)

_TITLE_FILL = PatternFill("solid", fgColor="D9D9D9")       # light grey (theme 2 approx)
_SECTION_FILL = PatternFill("solid", fgColor="BFBFBF")     # darker grey (section headers)
_HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")      # same as title
_COMMENTS_FILL = PatternFill("solid", fgColor="FFFF00")    # yellow for Comments column
_LABEL_FILL = PatternFill("solid", fgColor="F2F2F2")       # very light grey for label rows
_OLD_VERSION_FONT = Font(name="Times New Roman", size=11, italic=True, color="808080")
_OLD_VERSION_FILL = PatternFill("solid", fgColor="F5F5F5")  # faint grey for superseded rows

# ---------------------------------------------------------------------------
# Contract Analysis categories — order matches CR2A_Template.docx
# Maps cat_key → (section_name, display_label)
# ---------------------------------------------------------------------------
CONTRACT_ANALYSIS_SECTIONS = [
    ("I. Administrative & Commercial Terms", [
        ("contract_term_renewal_extensions", "Contract Term, Renewal & Extensions"),
        ("bonding_surety_insurance", "Bonding, Surety, & Insurance Obligations"),
        ("retainage_progress_payments", "Retainage, Progress Payments & Final Payment Terms"),
        ("pay_when_paid_if_paid", "Pay-When-Paid, Pay-If-Paid, or Owner Payment Contingencies"),
        ("price_escalation", "Price Escalation Clauses (Labor, Materials, Fuel, Inflation Adjustments)"),
        ("change_orders", "Change Orders, Scope Adjustments & Modifications"),
        ("termination_for_convenience", "Termination for Convenience (Owner/Agency Right to Terminate Without Cause)"),
        ("termination_for_cause", "Termination for Cause / Default by Contractor"),
        ("bid_protest", "Bid Protest Procedures & Claims of Improper Award"),
        ("bid_tabulation", "Bid Tabulation, Competition & Award Process Requirements"),
        ("contractor_qualification", "Contractor Qualification, Licensing & Certification Requirements"),
        ("release_orders", "Release Orders, Task Orders & Work Authorization Protocols"),
        ("assignment_novation", "Assignment & Novation Restrictions (Transfer of Contract Rights)"),
        ("audit_rights", "Audit Rights, Recordkeeping & Document Retention Obligations"),
        ("notice_requirements", "Notice Requirements & Claim Timeframes"),
    ]),
    ("II. Technical & Performance Terms", [
        ("scope_of_work", "Scope of Work (Work Inclusions, Exclusions & Defined Deliverables)"),
        ("performance_schedule", "Performance Schedule, Time for Completion & Critical Path Obligations"),
        ("delays", "Delays of Any Kind (Force Majeure, Acts of God, Weather, Owner-Caused)"),
        ("suspension_of_work", "Suspension of Work, Work Stoppages & Agency Directives"),
        ("submittals", "Submittals, Documentation & Approval Requirements"),
        ("emergency_work", "Emergency & Contingency Work Obligations"),
        ("permits_licensing", "Permits, Licensing & Regulatory Approvals for Work"),
        ("warranty", "Warranty, Guarantee & Defects Liability Periods"),
        ("use_of_aps_tools", "Use of APS Tools, Equipment, Materials or Supplies"),
        ("owner_supplied_support", "Owner-Supplied Support, Utilities & Site Access Provisions"),
        ("field_ticket", "Field Ticket, Daily Work Log & Documentation Requirements"),
        ("mobilization_demobilization", "Mobilization & Demobilization Provisions"),
        ("utility_coordination", "Utility Coordination, Locate Risk & Conflict Avoidance"),
        ("delivery_deadlines", "Delivery Deadlines, Milestone Dates, Substantial & Final Completion"),
        ("punch_list", "Punch List, Closeout Procedures & Acceptance of Work"),
        ("worksite_coordination", "Worksite Coordination, Access Restrictions & Sequencing Obligations"),
        ("deliverables", "Deliverables, Digital Submissions & Documentation Standards"),
    ]),
    ("III. Legal Risk & Enforcement", [
        ("indemnification", "Indemnification, Defense & Hold Harmless Provisions"),
        ("duty_to_defend", "Duty to Defend vs. Indemnify Scope Clarifications"),
        ("limitation_of_liability", "Limitations of Liability, Damage Caps & Waivers of Consequential Damages"),
        ("insurance_coverage", "Insurance Coverage, Additional Insured & Waiver of Subrogation Clauses"),
        ("dispute_resolution", "Dispute Resolution (Mediation, Arbitration, Litigation)"),
        ("flow_down_clauses", "Flow-Down Clauses (Prime-to-Subcontract Risk Pass-Through)"),
        ("subcontracting", "Subcontracting Restrictions, Approval & Substitution Requirements"),
        ("safety_osha", "Safety Standards, OSHA Compliance & Site-Specific Safety Obligations"),
        ("site_conditions", "Site Conditions, Differing Site Conditions & Changed Circumstances"),
        ("environmental", "Environmental Hazards, Waste Disposal & Hazardous Materials"),
        ("order_of_precedence", "Conflicting Documents / Order of Precedence Clauses"),
        ("setoff_withholding", "Setoff & Withholding Rights (Owner's Right to Deduct or Withhold Payment)"),
    ]),
    ("IV. Regulatory & Compliance Terms", [
        ("certified_payroll", "Certified Payroll, Recordkeeping & Reporting Obligations"),
        ("prevailing_wage", "Prevailing Wage, Davis-Bacon & Federal/State Wage Compliance"),
        ("eeo", "EEO, Non-Discrimination, MWBE/DBE Participation Requirements"),
        ("mwbe_dbe", "MWBE/DBE Participation Goals & Utilization Requirements"),
        ("apprenticeship", "Apprenticeship, Training & Workforce Development Requirements"),
        ("e_verify", "Immigration / E-Verify Compliance Obligations"),
        ("worker_classification", "Worker Classification & Independent Contractor Restrictions"),
        ("drug_free_workplace", "Drug-Free Workplace Programs & Substance Testing Requirements"),
    ]),
    ("V. Data, Technology & Deliverables", [
        ("data_ownership", "Data Ownership, Access & Rights to Digital Deliverables"),
        ("ai_technology_use", "AI / Technology Use Restrictions"),
        ("cybersecurity", "Cybersecurity Standards, Breach Notification & IT System Use Policies"),
        ("digital_deliverables", "Digital Deliverables, BIM/CAD Requirements & Electronic Submissions"),
        ("document_retention", "Document Retention, Records Preservation & Data Security Policies"),
        ("confidentiality", "Confidentiality, Data Security & Records Retention Obligations"),
    ]),
]

# ---------------------------------------------------------------------------
# Bid Checklist row mapping: item_key → row label in CC Bid Checklist
# This maps the bid_review_models field names to the Excel row labels.
# ---------------------------------------------------------------------------
BID_CHECKLIST_ROW_MAP = {
    # ProjectInformation — mapped to rows 3-17
    "project_title": "Bid Name",
    "solicitation_number": "Bid #",
    "owner": "Owner & Owner Address ",
    "scope": "Scope",
    "bid_model": "Bidding As:",
    # StandardContractItems — rows 36-55
    "pre_bid": "Pre-Bid?",
    "submission_format": "Submission type: Electronic / Hardcopy; Copies?",
    "bid_bond": "Bid Bond?",
    "payment_performance_bonds": "P & P Bonds?",
    "contract_time": "Contract Time",
    "liquidated_damages": "Liquidated Damages",
    "warranty": "Warranty?",
    "contractor_license": "License?",
    "insurance": "Insurance",
    "minority_dbe_goals": "MBE/DBE Goals?",
    "working_hours": "Working Hours",
    "subcontracting": "Subcontracting",
    "funding": "Funding",
    "certified_payroll": "Certified Payroll",
    "retainage": "Retainage",
    "safety": "Safety",
    "qualifications": "Qualifications",
    # SiteConditions — rows 56-64
    "site_access": "Site Access",
    "site_restoration": "Site Restoration",
    "bypass": "Bypass & Flow?",
    "traffic_control": "Traffic Control?",
    "disposal": "Disposal?",
    "water_hydrant_meter": "Hydrant Meter & Water?",
    # Cleaning — rows 92-95
    "cleaning_method": "Cleaning Method",
    "cleaning_passes": "Cleaning Passes",
    "cleaning_notifications": "Notifications?",  # row 95
    # CCTV — rows 98-100
    "nassco": "NASSCO",
    "cctv_submittal_format": "CCTV Submittal Format",
    # CIPP — rows 104-119
    "curing_method": "Curing Method",
    "cure_water": "Cure Water",
    "cipp_warranty": "Warranty",  # row 107
    "cipp_notifications": "Notifications?",  # row 108
    "contractor_qualifications": "Contractor Qualifications",
    "wet_out_facility": "Wet-Out Facility",
    "end_seals": "End Seals",
    "mudding_the_ends": "Mudding the Ends",
    "conditions_above_pipes": "Conditions Above Pipes/Overhead",
    "pre_liner": "Pre-Liner",
    "pipe_information": "Pipe Information",
    "resin_type": "Resin Type",
    "testing": "Testing",  # row 117
    "engineered_design_stamp": "Engineered Design Stamp Calculations",
    "air_testing": "Air Testing",
    # ManholeRehab — rows 124-141
    "mh_information": "MH Information",
    "mh_product_type": "Product Type",
    "products": "Products",
    "mh_testing": "Testing",  # row 128
    "mh_warranty": "Warranty",  # row 129
    "thickness": "Thickness",
    "compressive_strength": "Compressive Strength",
    "bond_strength": "Bond Strength",
    "shrinkage": "Shrinkage",
    "grout": "Grout",
    "measurement_payment": "Measurement & Payment Section",
    "external_coating": "External Coating of Manhole?",
    "mh_notifications": "Notifications?",  # row 137
    "nace": "NACE",
    "mh_bypass": "Bypass",  # row 139
    "substitution_requirements": "Substitution Requirements",
    "corrugations": "Corrugations",
}


def _get_template_path() -> Path:
    """Return path to the bundled Template.xlsx."""
    # When running from source: e:/CR2A/templates/Template.xlsx
    # When bundled with PyInstaller: look relative to the executable
    src_dir = Path(__file__).resolve().parent
    candidates = [
        src_dir.parent / "templates" / "Template.xlsx",   # source tree
        src_dir / "templates" / "Template.xlsx",           # alternate
        Path(os.environ.get("CR2A_TEMPLATE", "")) if os.environ.get("CR2A_TEMPLATE") else None,
    ]
    for p in candidates:
        if p and p.exists():
            return p
    raise FileNotFoundError("Template.xlsx not found in any expected location")


class ExcelTemplateBuilder:
    """
    Manages the per-bid-folder analysis workbook.

    On first use, copies Template.xlsx into the bid folder, adds the
    Contract Analysis and Specs sheets, then provides update_*() methods
    that the GUI callbacks invoke to populate data as analyses complete.
    """

    SHEET_CONTRACT = "Contract Analysis"
    SHEET_SPECS = "Specs"
    SHEET_BID = "CC Bid Checklist"

    def __init__(self, project_root: Path, contract_files: Optional[List[str]] = None):
        """
        Args:
            project_root: The bid folder path (e.g., F:\\Bids\\Gateway).
            contract_files: List of contract filenames in the folder.
        """
        self.project_root = Path(project_root)
        self.contract_files = contract_files or []
        self._excel_path: Optional[Path] = None
        # Row lookup caches
        self._contract_row_map: Dict[str, int] = {}   # cat_key → row
        self._bid_row_map: Dict[str, int] = {}          # label → row

    WORKBOOK_NAME = "CR2A_Analysis.xlsx"

    @property
    def excel_path(self) -> Path:
        """Path to the workbook in the bid folder or a parent folder."""
        if self._excel_path and self._excel_path.exists():
            return self._excel_path
        # Default name
        return self.project_root / self.WORKBOOK_NAME

    def _find_existing_workbook(self) -> Optional[Path]:
        """Search project_root and its parent directories for an existing workbook.

        Walks up to 3 levels above project_root to find a CR2A_Analysis.xlsx
        that was created by a previous analysis. Returns None if not found.
        """
        # Check project_root first
        candidate = self.project_root / self.WORKBOOK_NAME
        if candidate.exists():
            return candidate

        # Walk up parent directories (max 3 levels)
        current = self.project_root.parent
        for _ in range(3):
            candidate = current / self.WORKBOOK_NAME
            if candidate.exists():
                logger.info("Found existing workbook in parent folder: %s", candidate)
                return candidate
            parent = current.parent
            if parent == current:
                break  # reached filesystem root
            current = parent

        return None

    # ------------------------------------------------------------------
    # Initialization
    # ------------------------------------------------------------------

    def initialize_workbook(self) -> Path:
        """
        Find or create the workbook, add Contract Analysis
        and Specs sheets, return the workbook path.

        Searches parent directories for an existing workbook before
        creating a new one from the template.
        """
        dest = self._find_existing_workbook()
        if dest:
            logger.info("Using existing workbook: %s", dest)
        else:
            dest = self.project_root / self.WORKBOOK_NAME
            template = _get_template_path()
            shutil.copy2(template, dest)
            logger.info("Copied template to %s", dest)

        self._excel_path = dest

        # Ensure the new sheets exist
        wb = openpyxl.load_workbook(dest)
        changed = False
        if self.SHEET_CONTRACT not in wb.sheetnames:
            self._create_contract_analysis_sheet(wb)
            changed = True
        else:
            # Migrate existing sheet: add versioning columns if missing
            if self._migrate_contract_sheet(wb):
                changed = True
        if self.SHEET_SPECS not in wb.sheetnames:
            self._create_specs_sheet(wb)
            changed = True
        if changed:
            wb.save(dest)
            logger.info("Added analysis sheets to %s", dest)
        wb.close()

        # Build row caches
        self._build_row_caches()
        return dest

    # ------------------------------------------------------------------
    # Sheet creation
    # ------------------------------------------------------------------

    def _create_contract_analysis_sheet(self, wb: openpyxl.Workbook) -> None:
        """Create the Contract Analysis sheet with all sections/items."""
        ws = wb.create_sheet(self.SHEET_CONTRACT)

        # Column widths matching CC Bid Checklist
        ws.column_dimensions["A"].width = 20.55
        ws.column_dimensions["B"].width = 89.82
        ws.column_dimensions["C"].width = 37.27
        ws.column_dimensions["D"].width = 15.0   # hidden: contract path
        ws.column_dimensions["E"].width = 12.0   # Open link
        ws.column_dimensions["F"].width = 10.0   # Version
        ws.column_dimensions["G"].width = 18.0   # Analyzed Date
        ws.column_dimensions["H"].width = 10.0   # hidden: cat_key identifier

        # Row 1: Title
        ws.merge_cells("A1:H1")
        c = ws["A1"]
        c.value = "Contract Clause Risk & Compliance Summary"
        c.font = _TITLE_FONT
        c.fill = _TITLE_FILL
        c.alignment = _CENTER_WRAP
        c.border = _THIN_BORDER

        # Row 2: Column headers
        headers = [
            ("A2", "Clause Category", _HEADER_FILL),
            ("B2", "Clause Summary / Location", _HEADER_FILL),
            ("C2", "Redline Recommendations", _COMMENTS_FILL),
            ("D2", "File Path", _HEADER_FILL),
            ("E2", "Source", _HEADER_FILL),
            ("F2", "Version", _HEADER_FILL),
            ("G2", "Analyzed", _HEADER_FILL),
            ("H2", "cat_key", _HEADER_FILL),
        ]
        for cell_ref, label, fill in headers:
            c = ws[cell_ref]
            c.value = label
            c.font = _HEADER_FONT
            c.fill = fill
            c.alignment = _CENTER_WRAP
            c.border = _THIN_BORDER

        # Freeze panes at row 3
        ws.freeze_panes = "A3"

        # Build rows from CONTRACT_ANALYSIS_SECTIONS
        row = 3
        for section_title, items in CONTRACT_ANALYSIS_SECTIONS:
            # Section header row (merged)
            ws.merge_cells(f"A{row}:H{row}")
            c = ws[f"A{row}"]
            c.value = section_title
            c.font = _TITLE_FONT
            c.fill = _SECTION_FILL
            c.alignment = _CENTER_WRAP
            c.border = _THIN_BORDER
            row += 1

            for cat_key, display_label in items:
                # Col A: label
                ws[f"A{row}"].value = display_label
                ws[f"A{row}"].font = _LABEL_FONT
                ws[f"A{row}"].fill = _LABEL_FILL
                ws[f"A{row}"].alignment = _TOP_WRAP
                ws[f"A{row}"].border = _THIN_BORDER

                # Col B: summary (empty, filled by analysis)
                ws[f"B{row}"].font = _DATA_FONT
                ws[f"B{row}"].alignment = _TOP_WRAP
                ws[f"B{row}"].border = _THIN_BORDER

                # Col C: redline/comments (empty)
                ws[f"C{row}"].font = _DATA_FONT
                ws[f"C{row}"].alignment = _TOP_WRAP
                ws[f"C{row}"].border = _THIN_BORDER

                # Col D: file path (hidden helper for HYPERLINK)
                ws[f"D{row}"].font = _DATA_FONT
                ws[f"D{row}"].border = _THIN_BORDER

                # Col E: HYPERLINK formula (populated when data is written)
                ws[f"E{row}"].font = _DATA_FONT
                ws[f"E{row}"].border = _THIN_BORDER

                # Col F: Version (empty until first analysis)
                ws[f"F{row}"].font = _DATA_FONT
                ws[f"F{row}"].alignment = _CENTER_WRAP
                ws[f"F{row}"].border = _THIN_BORDER

                # Col G: Analyzed date (empty until first analysis)
                ws[f"G{row}"].font = _DATA_FONT
                ws[f"G{row}"].alignment = _CENTER_WRAP
                ws[f"G{row}"].border = _THIN_BORDER

                # Col H: cat_key identifier (for row cache rebuild)
                ws[f"H{row}"].value = cat_key
                ws[f"H{row}"].font = _DATA_FONT
                ws[f"H{row}"].border = _THIN_BORDER

                ws.row_dimensions[row].height = 30.0
                row += 1

            # Spacer row after each section
            row += 1

        # Hide helper columns
        ws.column_dimensions["D"].hidden = True
        ws.column_dimensions["H"].hidden = True

    def _create_specs_sheet(self, wb: openpyxl.Workbook) -> None:
        """Create the Specs sheet for technical specification extraction."""
        ws = wb.create_sheet(self.SHEET_SPECS)

        ws.column_dimensions["A"].width = 30.0
        ws.column_dimensions["B"].width = 70.0
        ws.column_dimensions["C"].width = 30.0

        # Row 1: Title
        ws.merge_cells("A1:C1")
        c = ws["A1"]
        c.value = "Technical Specification Extraction"
        c.font = _TITLE_FONT
        c.fill = _TITLE_FILL
        c.alignment = _CENTER_WRAP
        c.border = _THIN_BORDER

        # Row 2: Headers
        for cell_ref, label, fill in [
            ("A2", "Specification", _HEADER_FILL),
            ("B2", "Required Value / Requirement", _HEADER_FILL),
            ("C2", "Source Location", _COMMENTS_FILL),
        ]:
            c = ws[cell_ref]
            c.value = label
            c.font = _HEADER_FONT
            c.fill = fill
            c.alignment = _CENTER_WRAP
            c.border = _THIN_BORDER

        ws.freeze_panes = "A3"

    # ------------------------------------------------------------------
    # Migration for pre-versioning workbooks
    # ------------------------------------------------------------------

    def _migrate_contract_sheet(self, wb: openpyxl.Workbook) -> bool:
        """Add versioning columns (F, G, H) to an existing Contract Analysis sheet.

        Returns True if the sheet was modified (needs save).
        """
        ws = wb[self.SHEET_CONTRACT]
        # Check if column H header already exists
        if ws["H2"].value == "cat_key":
            return False  # already migrated

        logger.info("Migrating Contract Analysis sheet to add versioning columns")

        # Set column widths
        ws.column_dimensions["F"].width = 10.0
        ws.column_dimensions["G"].width = 18.0
        ws.column_dimensions["H"].width = 10.0

        # Add headers in row 2
        for cell_ref, label, fill in [
            ("F2", "Version", _HEADER_FILL),
            ("G2", "Analyzed", _HEADER_FILL),
            ("H2", "cat_key", _HEADER_FILL),
        ]:
            c = ws[cell_ref]
            c.value = label
            c.font = _HEADER_FONT
            c.fill = fill
            c.alignment = _CENTER_WRAP
            c.border = _THIN_BORDER

        # Walk the deterministic layout to stamp cat_key into column H
        row = 3
        for _section_title, items in CONTRACT_ANALYSIS_SECTIONS:
            row += 1  # section header
            for cat_key, _label in items:
                ws[f"H{row}"].value = cat_key
                ws[f"H{row}"].font = _DATA_FONT
                ws[f"H{row}"].border = _THIN_BORDER
                ws[f"F{row}"].font = _DATA_FONT
                ws[f"F{row}"].alignment = _CENTER_WRAP
                ws[f"F{row}"].border = _THIN_BORDER
                ws[f"G{row}"].font = _DATA_FONT
                ws[f"G{row}"].alignment = _CENTER_WRAP
                ws[f"G{row}"].border = _THIN_BORDER
                row += 1
            row += 1  # spacer

        ws.column_dimensions["H"].hidden = True

        # Expand title merge to include new columns
        try:
            ws.unmerge_cells("A1:E1")
        except Exception:
            pass
        ws.merge_cells("A1:H1")

        # Expand section header merges
        row = 3
        for _section_title, items in CONTRACT_ANALYSIS_SECTIONS:
            try:
                ws.unmerge_cells(f"A{row}:E{row}")
            except Exception:
                pass
            ws.merge_cells(f"A{row}:H{row}")
            row += 1
            row += len(items)
            row += 1  # spacer

        return True

    # ------------------------------------------------------------------
    # Row cache building
    # ------------------------------------------------------------------

    def _build_row_caches(self) -> None:
        """Scan sheets to build cat_key → row and label → row maps.

        For Contract Analysis, scans column H (cat_key identifier) so the
        cache stays correct even after version-history rows are inserted.
        Only the *first* row for each cat_key is cached — that is always
        the current-version row.
        """
        try:
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        except Exception as e:
            logger.warning("Could not build row caches: %s", e)
            return

        # Contract Analysis: scan column H for cat_key identifiers.
        # Skip prefixed keys: ~cat_key = version history, +cat_key = sub-instance
        self._contract_row_map.clear()
        if self.SHEET_CONTRACT in wb.sheetnames:
            ws = wb[self.SHEET_CONTRACT]
            for row in range(3, ws.max_row + 1):
                val = ws[f"H{row}"].value
                if val and isinstance(val, str):
                    key = val.strip()
                    if key and not key.startswith(("~", "+")) and key not in self._contract_row_map:
                        self._contract_row_map[key] = row

        # CC Bid Checklist: scan column A labels
        self._bid_row_map.clear()
        if self.SHEET_BID in wb.sheetnames:
            ws = wb[self.SHEET_BID]
            for row in range(1, ws.max_row + 1):
                val = ws[f"A{row}"].value
                if val and isinstance(val, str):
                    label = val.strip()
                    if label:
                        self._bid_row_map[label] = row

        wb.close()

    # ------------------------------------------------------------------
    # Data update methods
    # ------------------------------------------------------------------

    def update_contract_category(self, cat_key: str, clause_block: dict,
                                 contract_file: str = "") -> bool:
        """
        Write a single contract analysis result to the Contract Analysis sheet.

        If the category already has data from a prior analysis, the old data
        is pushed to a version-history row directly below and the new result
        becomes the current version.  Each row carries a version number and
        timestamp so reviewers can see the full analysis history.

        Args:
            cat_key: Category key (e.g., "change_orders").
            clause_block: Dict with keys like 'Clause Location', 'Clause Summary',
                         'Redline Recommendations', 'Harmful Language / Policy Conflicts'.
            contract_file: Filename of the source contract for HYPERLINK.

        Returns:
            True if write succeeded, False if file was locked or error.
        """
        row = self._contract_row_map.get(cat_key)
        if not row:
            logger.warning("No row mapping for contract category: %s", cat_key)
            return False

        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb[self.SHEET_CONTRACT]

            # ----------------------------------------------------------
            # Versioning: detect whether this row already has data
            # ----------------------------------------------------------
            existing_summary = ws[f"B{row}"].value
            current_version = 1
            old_version_str = ws[f"F{row}"].value
            if old_version_str and isinstance(old_version_str, str):
                # Parse "v3 (Current)" → 3
                import re as _re
                m = _re.search(r'v(\d+)', old_version_str)
                if m:
                    current_version = int(m.group(1))

            if existing_summary:
                # There is prior data — archive it to a history row below.
                # First, remove any old sub-instance rows (+cat_key) from
                # the prior analysis — they'll be regenerated by
                # update_contract_category_multi().
                purge_row = row + 1
                while purge_row <= ws.max_row:
                    if ws[f"H{purge_row}"].value == f"+{cat_key}":
                        ws.delete_rows(purge_row)
                    else:
                        break

                # Insert directly after the current row so history reads
                # newest-first (reverse chronological) top to bottom.
                # Temporarily unmerge ranges below to avoid openpyxl merge shift bugs.
                insert_at = row + 1
                saved_merges = []
                for m in list(ws.merged_cells.ranges):
                    if m.min_row >= insert_at:
                        saved_merges.append((m.min_row, m.max_row, m.min_col, m.max_col))
                        ws.unmerge_cells(str(m))

                ws.insert_rows(insert_at)

                # Re-apply merges shifted by 1 row
                for min_r, max_r, min_c, max_c in saved_merges:
                    ws.merge_cells(
                        start_row=min_r + 1, end_row=max_r + 1,
                        start_column=min_c, end_column=max_c,
                    )

                # Copy old data into the history row with old-version styling
                for col_letter in ("B", "C", "D", "E", "G"):
                    old_val = ws[f"{col_letter}{row}"].value
                    ws[f"{col_letter}{insert_at}"].value = old_val
                    ws[f"{col_letter}{insert_at}"].font = _OLD_VERSION_FONT
                    ws[f"{col_letter}{insert_at}"].fill = _OLD_VERSION_FILL
                    ws[f"{col_letter}{insert_at}"].alignment = _TOP_WRAP
                    ws[f"{col_letter}{insert_at}"].border = _THIN_BORDER

                # Mark the archived row
                ws[f"A{insert_at}"].value = ""  # no label — indent under parent
                ws[f"A{insert_at}"].font = _OLD_VERSION_FONT
                ws[f"A{insert_at}"].fill = _OLD_VERSION_FILL
                ws[f"A{insert_at}"].border = _THIN_BORDER
                ws[f"F{insert_at}"].value = f"v{current_version}"
                ws[f"F{insert_at}"].font = _OLD_VERSION_FONT
                ws[f"F{insert_at}"].fill = _OLD_VERSION_FILL
                ws[f"F{insert_at}"].alignment = _CENTER_WRAP
                ws[f"F{insert_at}"].border = _THIN_BORDER
                # Tag history rows with ~cat_key so we can skip them in scans
                ws[f"H{insert_at}"].value = f"~{cat_key}"
                ws[f"H{insert_at}"].font = _OLD_VERSION_FONT
                ws[f"H{insert_at}"].border = _THIN_BORDER

                ws.row_dimensions[insert_at].height = 25.0

                current_version += 1
                logger.info("Archived v%d for %s at row %d",
                            current_version - 1, cat_key, insert_at)

            # ----------------------------------------------------------
            # Write new / current data to the primary row
            # ----------------------------------------------------------
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

            # Column B: Summary + Location
            summary = clause_block.get("Clause Summary", "")
            location = clause_block.get("Clause Location", "")
            page = clause_block.get("Clause Page", "")
            parts = []
            if location:
                parts.append(f"[{location}]")
            if page:
                parts.append(f"(Page {page})")
            if summary:
                parts.append(summary)
            ws[f"B{row}"].value = " ".join(parts) if parts else ""
            ws[f"B{row}"].font = _DATA_FONT
            ws[f"B{row}"].alignment = _TOP_WRAP

            # Column C: Redline + Harmful Language
            redline_parts = []
            redlines = clause_block.get("Redline Recommendations", [])
            if isinstance(redlines, list):
                for r in redlines:
                    if isinstance(r, dict):
                        action = r.get("Action", "")
                        text = r.get("Text", "")
                        if action or text:
                            redline_parts.append(f"{action}: {text}" if action else text)
                    elif isinstance(r, str):
                        redline_parts.append(r)
            harmful = clause_block.get("Harmful Language / Policy Conflicts", [])
            if harmful:
                if isinstance(harmful, list):
                    for h in harmful:
                        redline_parts.append(f"[HARMFUL] {h}")
                elif isinstance(harmful, str):
                    redline_parts.append(f"[HARMFUL] {harmful}")
            ws[f"C{row}"].value = "\n".join(redline_parts) if redline_parts else ""
            ws[f"C{row}"].font = _DATA_FONT
            ws[f"C{row}"].alignment = _TOP_WRAP

            # Column D: relative file path for HYPERLINK (with PDF coordinates)
            file_ref = contract_file or (self.contract_files[0] if self.contract_files else "")
            if file_ref:
                # Build a hyperlink with exact PDF coordinates when possible
                pdf_fragment = ""
                clause_location_text = clause_block.get("Clause Location", "")
                clause_page = clause_block.get("Clause Page")
                if file_ref.lower().endswith(".pdf") and clause_location_text:
                    try:
                        from contract_uploader import find_text_in_pdf
                        full_pdf_path = str(self.project_root / file_ref)
                        page_num, y_coord = find_text_in_pdf(
                            full_pdf_path, clause_location_text,
                            page_hint=clause_page
                        )
                        if page_num and y_coord is not None:
                            pdf_fragment = f"#page={page_num}&zoom=100,0,{y_coord:.0f}"
                        elif page_num:
                            pdf_fragment = f"#page={page_num}"
                        elif clause_page:
                            pdf_fragment = f"#page={clause_page}"
                    except Exception as e:
                        logger.debug("Could not resolve PDF coordinates: %s", e)
                        if clause_page:
                            pdf_fragment = f"#page={clause_page}"
                elif clause_page:
                    pdf_fragment = f"#page={clause_page}"

                ws[f"D{row}"].value = f".\\{file_ref}{pdf_fragment}"
                ws[f"E{row}"].value = f'=HYPERLINK(D{row},"Open Contract")'
                ws[f"E{row}"].font = Font(name="Times New Roman", size=11,
                                          color="0563C1", underline="single")

            # Column F: Version label
            ws[f"F{row}"].value = f"v{current_version} (Current)"
            ws[f"F{row}"].font = Font(name="Times New Roman", size=11, bold=True)
            ws[f"F{row}"].alignment = _CENTER_WRAP
            ws[f"F{row}"].border = _THIN_BORDER

            # Column G: Timestamp
            ws[f"G{row}"].value = timestamp
            ws[f"G{row}"].font = _DATA_FONT
            ws[f"G{row}"].alignment = _CENTER_WRAP
            ws[f"G{row}"].border = _THIN_BORDER

            wb.save(self.excel_path)
            wb.close()

            # Rebuild row caches — row insertion shifts all rows below
            if existing_summary:
                self._build_row_caches()

            return True

        except PermissionError:
            logger.warning("Workbook is open in another application. Close it to update.")
            return False
        except Exception as e:
            logger.error("Failed to update contract category %s: %s", cat_key, e)
            return False

    # ------------------------------------------------------------------
    # Multi-instance support
    # ------------------------------------------------------------------

    _MULTI_INSTANCE_FONT = Font(name="Times New Roman", size=11, italic=True)
    _MULTI_INSTANCE_FILL = PatternFill("solid", fgColor="EBF5FB")  # light blue tint

    def update_contract_category_multi(
        self, cat_key: str, clause_blocks: List[dict], contract_file: str = ""
    ) -> bool:
        """Write one or more clause instances for a single category.

        The first block is written to the primary row (via the normal
        versioning path).  Additional blocks are inserted as sub-rows
        directly below the primary row, tagged with ``+cat_key`` in
        column H so they can be distinguished from version-history rows
        (``~cat_key``).

        Args:
            cat_key:        Category key (e.g., "notice_requirements").
            clause_blocks:  List of clause_block dicts — one per distinct
                            clause instance found by the AI.
            contract_file:  Filename of the source contract for HYPERLINK.

        Returns:
            True if at least the primary write succeeded.
        """
        if not clause_blocks:
            return False

        # Write the primary instance through the normal versioned path
        primary = clause_blocks[0]
        ok = self.update_contract_category(cat_key, primary, contract_file)
        if not ok:
            return False

        # Additional instances (2nd, 3rd, …) — insert sub-rows
        extras = clause_blocks[1:]
        if not extras:
            return True

        row = self._contract_row_map.get(cat_key)
        if not row:
            return True  # primary succeeded; extras can't be placed

        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb[self.SHEET_CONTRACT]
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

            # Remove any existing sub-instance rows from a prior analysis
            # (they carry ``+cat_key`` in column H)
            purge_row = row + 1
            while purge_row <= ws.max_row:
                if ws[f"H{purge_row}"].value == f"+{cat_key}":
                    ws.delete_rows(purge_row)
                    # don't increment — next row shifted up
                else:
                    break  # sub-instance rows are contiguous

            # Insert new sub-instance rows right below primary.
            # openpyxl's insert_rows doesn't always shift merge ranges
            # correctly, so we temporarily unmerge any ranges in the
            # zone below the insert point, then re-merge after inserts.
            insert_pos = row + 1
            file_ref = contract_file or (self.contract_files[0] if self.contract_files else "")

            num_extras = len(extras)

            # Collect and remove merges that could collide with inserts
            saved_merges = []
            for m in list(ws.merged_cells.ranges):
                if m.min_row >= insert_pos:
                    saved_merges.append((m.min_row, m.max_row, m.min_col, m.max_col))
                    ws.unmerge_cells(str(m))

            # Process extras in reverse so they end up in the correct order
            for idx, block in zip(
                range(num_extras + 1, 1, -1),
                reversed(extras),
            ):
                ws.insert_rows(insert_pos)

                summary = block.get("Clause Summary", "")
                location = block.get("Clause Location", "")
                page = block.get("Clause Page", "")
                parts = []
                if location:
                    parts.append(f"[{location}]")
                if page:
                    parts.append(f"(Page {page})")
                if summary:
                    parts.append(summary)

                # Col A: instance label
                ws[f"A{insert_pos}"].value = f"  Instance {idx}"
                ws[f"A{insert_pos}"].font = self._MULTI_INSTANCE_FONT
                ws[f"A{insert_pos}"].fill = self._MULTI_INSTANCE_FILL
                ws[f"A{insert_pos}"].alignment = _TOP_WRAP
                ws[f"A{insert_pos}"].border = _THIN_BORDER

                # Col B: summary
                ws[f"B{insert_pos}"].value = " ".join(parts) if parts else ""
                ws[f"B{insert_pos}"].font = _DATA_FONT
                ws[f"B{insert_pos}"].fill = self._MULTI_INSTANCE_FILL
                ws[f"B{insert_pos}"].alignment = _TOP_WRAP
                ws[f"B{insert_pos}"].border = _THIN_BORDER

                # Col C: redlines
                redline_parts = []
                redlines = block.get("Redline Recommendations", [])
                if isinstance(redlines, list):
                    for r in redlines:
                        if isinstance(r, dict):
                            action = r.get("Action", "")
                            text = r.get("Text", "")
                            redline_parts.append(f"{action}: {text}" if action else text)
                        elif isinstance(r, str):
                            redline_parts.append(r)
                ws[f"C{insert_pos}"].value = "\n".join(redline_parts) if redline_parts else ""
                ws[f"C{insert_pos}"].font = _DATA_FONT
                ws[f"C{insert_pos}"].fill = self._MULTI_INSTANCE_FILL
                ws[f"C{insert_pos}"].alignment = _TOP_WRAP
                ws[f"C{insert_pos}"].border = _THIN_BORDER

                # Col D+E: hyperlink
                if file_ref:
                    pdf_fragment = ""
                    clause_location_text = block.get("Clause Location", "")
                    clause_page = block.get("Clause Page")
                    if file_ref.lower().endswith(".pdf") and clause_location_text:
                        try:
                            from contract_uploader import find_text_in_pdf
                            full_pdf_path = str(self.project_root / file_ref)
                            page_num, y_coord = find_text_in_pdf(
                                full_pdf_path, clause_location_text,
                                page_hint=clause_page
                            )
                            if page_num and y_coord is not None:
                                pdf_fragment = f"#page={page_num}&zoom=100,0,{y_coord:.0f}"
                            elif page_num:
                                pdf_fragment = f"#page={page_num}"
                            elif clause_page:
                                pdf_fragment = f"#page={clause_page}"
                        except Exception:
                            if clause_page:
                                pdf_fragment = f"#page={clause_page}"
                    elif clause_page:
                        pdf_fragment = f"#page={clause_page}"

                    ws[f"D{insert_pos}"].value = f".\\{file_ref}{pdf_fragment}"
                    ws[f"E{insert_pos}"].value = f'=HYPERLINK(D{insert_pos},"Open Contract")'
                    ws[f"E{insert_pos}"].font = Font(name="Times New Roman", size=11,
                                                     color="0563C1", underline="single")
                    ws[f"E{insert_pos}"].fill = self._MULTI_INSTANCE_FILL
                    ws[f"E{insert_pos}"].border = _THIN_BORDER

                # Col F: no version — these are sub-instances, not versions
                ws[f"F{insert_pos}"].font = _DATA_FONT
                ws[f"F{insert_pos}"].fill = self._MULTI_INSTANCE_FILL
                ws[f"F{insert_pos}"].alignment = _CENTER_WRAP
                ws[f"F{insert_pos}"].border = _THIN_BORDER

                # Col G: timestamp
                ws[f"G{insert_pos}"].value = timestamp
                ws[f"G{insert_pos}"].font = _DATA_FONT
                ws[f"G{insert_pos}"].fill = self._MULTI_INSTANCE_FILL
                ws[f"G{insert_pos}"].alignment = _CENTER_WRAP
                ws[f"G{insert_pos}"].border = _THIN_BORDER

                # Col H: tag as sub-instance
                ws[f"H{insert_pos}"].value = f"+{cat_key}"
                ws[f"H{insert_pos}"].font = _DATA_FONT
                ws[f"H{insert_pos}"].border = _THIN_BORDER

                ws.row_dimensions[insert_pos].height = 30.0

            # Re-apply saved merges, shifted down by the number of inserted rows
            for min_r, max_r, min_c, max_c in saved_merges:
                new_min = min_r + num_extras
                new_max = max_r + num_extras
                ws.merge_cells(
                    start_row=new_min, end_row=new_max,
                    start_column=min_c, end_column=max_c,
                )

            wb.save(self.excel_path)
            wb.close()
            self._build_row_caches()

            logger.info("Wrote %d instance(s) for %s", len(clause_blocks), cat_key)
            return True

        except PermissionError:
            logger.warning("Workbook is open in another application. Close it to update.")
            return True  # primary succeeded
        except Exception as e:
            logger.error("Failed to write multi-instance for %s: %s", cat_key, e)
            return True  # primary succeeded

    def update_bid_review_item(self, item_key: str, item: Any,
                               section_key: str = "",
                               contract_file: str = "") -> bool:
        """
        Write a single bid review item to the CC Bid Checklist sheet.

        Args:
            item_key: Field name (e.g., "retainage", "bid_bond").
            item: A ChecklistItem (or dict with value/location/confidence/notes/page).
            section_key: The section this item belongs to (for disambiguation).
            contract_file: Filename of the source contract for HYPERLINK.

        Returns:
            True if write succeeded.
        """
        # Resolve the Excel row label for this item
        row_label = BID_CHECKLIST_ROW_MAP.get(item_key)
        if not row_label:
            logger.debug("No row mapping for bid item: %s", item_key)
            return False

        row = self._bid_row_map.get(row_label)
        if not row:
            # Try stripping whitespace from keys
            for label, r in self._bid_row_map.items():
                if label.strip() == row_label.strip():
                    row = r
                    break
        if not row:
            logger.debug("Row label '%s' not found in checklist", row_label)
            return False

        # Extract values from ChecklistItem or dict
        if hasattr(item, "value"):
            value = item.value or ""
            location = item.location or ""
            confidence = item.confidence or ""
            notes = item.notes or ""
            page = item.page or ""
        elif isinstance(item, dict):
            value = item.get("value", "")
            location = item.get("location", "")
            confidence = item.get("confidence", "")
            notes = item.get("notes", "")
            page = item.get("page", "")
        else:
            return False

        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb[self.SHEET_BID]

            # Column B (Notes): value + location + page
            parts = []
            if value:
                parts.append(str(value))
            if location:
                parts.append(f"[{location}]")
            if page:
                parts.append(f"(Page {page})")
            ws[f"B{row}"].value = " ".join(parts) if parts else ""

            # Column C (Comments): confidence + notes
            comment_parts = []
            if confidence and confidence != "not_found":
                comment_parts.append(f"Confidence: {confidence}")
            if notes:
                comment_parts.append(str(notes))
            ws[f"C{row}"].value = " | ".join(comment_parts) if comment_parts else ""

            wb.save(self.excel_path)
            wb.close()
            return True

        except PermissionError:
            logger.warning("Workbook is open in another application. Close it to update.")
            return False
        except Exception as e:
            logger.error("Failed to update bid item %s: %s", item_key, e)
            return False

    def update_bid_review_full(self, result: Any) -> int:
        """
        Write all bid review results at once.

        Args:
            result: A BidChecklistResult object.

        Returns:
            Number of items successfully written.
        """
        count = 0
        if result is None:
            return 0

        # Walk all sections and their items
        section_attrs = [
            ("project_information", "project_information"),
            ("standard_contract_items", "standard_contract_items"),
            ("site_conditions", "site_conditions"),
            ("cleaning", "cleaning"),
            ("cctv", "cctv"),
            ("cipp", "cipp"),
            ("manhole_rehab", "manhole_rehab"),
        ]
        for section_key, attr_name in section_attrs:
            section = getattr(result, attr_name, None)
            if section is None:
                continue
            field_map = getattr(section, "FIELD_MAP", {})
            for display_name, field_name in field_map.items():
                item = getattr(section, field_name, None)
                if item is None:
                    continue
                # Build a disambiguated key for sections with duplicate labels
                excel_key = self._disambiguate_item_key(field_name, section_key)
                if self.update_bid_review_item(excel_key, item, section_key=section_key):
                    count += 1

            # Handle nested sub-sections (CIPP design_requirements, MH spincast)
            if hasattr(section, "design_requirements") and section.design_requirements:
                sub = section.design_requirements
                sub_map = getattr(sub, "FIELD_MAP", {})
                for display_name, field_name in sub_map.items():
                    item = getattr(sub, field_name, None)
                    if item:
                        # These don't have rows in the current template yet
                        logger.debug("CIPP design req '%s' — no template row", field_name)

            if hasattr(section, "spincast") and section.spincast:
                sub = section.spincast
                sub_map = getattr(sub, "FIELD_MAP", {})
                for display_name, field_name in sub_map.items():
                    item = getattr(sub, field_name, None)
                    if item:
                        logger.debug("Spincast '%s' — no template row", field_name)

        return count

    def _disambiguate_item_key(self, field_name: str, section_key: str) -> str:
        """
        Some field names appear in multiple sections (e.g., 'warranty',
        'notifications', 'testing'). Return a unique key for the
        BID_CHECKLIST_ROW_MAP lookup.
        """
        # Check if the base field_name is in the map
        if field_name in BID_CHECKLIST_ROW_MAP:
            return field_name

        # Try section-prefixed versions
        section_prefixes = {
            "cipp": "cipp_",
            "manhole_rehab": "mh_",
            "cleaning": "cleaning_",
        }
        prefix = section_prefixes.get(section_key, "")
        prefixed = f"{prefix}{field_name}"
        if prefixed in BID_CHECKLIST_ROW_MAP:
            return prefixed

        return field_name

    def update_specs(self, specs_text: str) -> bool:
        """
        Write specs extraction results to the Specs sheet.

        Args:
            specs_text: The raw text output from specs analysis.

        Returns:
            True if write succeeded.
        """
        if not specs_text:
            return False

        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb[self.SHEET_SPECS]

            # Clear existing data rows (keep headers)
            for row in range(3, ws.max_row + 1):
                for col in range(1, 4):
                    ws.cell(row=row, column=col).value = None

            # Parse specs text into rows — split on newlines,
            # try to detect "label: value" patterns
            row = 3
            for line in specs_text.strip().split("\n"):
                line = line.strip()
                if not line:
                    continue

                if ":" in line:
                    parts = line.split(":", 1)
                    ws[f"A{row}"].value = parts[0].strip()
                    ws[f"A{row}"].font = _LABEL_FONT
                    ws[f"A{row}"].alignment = _TOP_WRAP
                    ws[f"A{row}"].border = _THIN_BORDER
                    ws[f"B{row}"].value = parts[1].strip()
                    ws[f"B{row}"].font = _DATA_FONT
                    ws[f"B{row}"].alignment = _TOP_WRAP
                    ws[f"B{row}"].border = _THIN_BORDER
                else:
                    # Could be a section header or freeform text
                    ws[f"A{row}"].value = line
                    ws[f"A{row}"].font = _LABEL_FONT
                    ws[f"A{row}"].alignment = _TOP_WRAP
                    ws[f"A{row}"].border = _THIN_BORDER

                ws.row_dimensions[row].height = 20.0
                row += 1

            wb.save(self.excel_path)
            wb.close()
            return True

        except PermissionError:
            logger.warning("Workbook is open. Close it to update specs.")
            return False
        except Exception as e:
            logger.error("Failed to update specs: %s", e)
            return False

    def update_title(self, project_name: str = "") -> bool:
        """Update the title row of CC Bid Checklist with the project name."""
        if not project_name:
            return False
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb[self.SHEET_BID]
            ws["A1"].value = project_name
            wb.save(self.excel_path)
            wb.close()
            return True
        except PermissionError:
            return False
        except Exception as e:
            logger.error("Failed to update title: %s", e)
            return False
